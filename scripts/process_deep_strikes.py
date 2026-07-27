#!/usr/bin/env python3

"""
Ukraine–Russia Deep Strike Monitor
Excel -> JSON feldolgozó

Bemeneti Excel:
data/manual/ukrajna_oroszorszag_melysegi_csapasok.xlsx

Kötelező munkalapok:
- Ukrajna → Oroszország
- Oroszország → Ukrajna

Kimenetek:
- data/deep_strikes.json
- data/deep_strikes_summary.json
- data/deep_strikes_validation.json

A script visszafelé kompatibilis a jelenlegi Excellel.

Az alap magyar oszlopok továbbra is kötelezőek.
Az angol és részletes leírás mezők opcionálisak.

Támogatott opcionális Excel-oszlopok:
- Régió EN
- Helyszín EN
- Támadás típusa EN
- Célpont típusa EN
- Rövid leírás EN
- Részletes leírás HU
- Részletes leírás EN
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys

from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ============================================================
# BEÁLLÍTÁSOK
# ============================================================

DEFAULT_INPUT = (
    "data/manual/ukrajna_oroszorszag_melysegi_csapasok.xlsx"
)

DEFAULT_OUTPUT = "data/deep_strikes.json"
DEFAULT_SUMMARY = "data/deep_strikes_summary.json"
DEFAULT_VALIDATION = "data/deep_strikes_validation.json"


SHEETS = {
    "Ukrajna → Oroszország": "UA_RU",
    "Oroszország → Ukrajna": "RU_UA",
}


# A jelenlegi Excel oszlopai.
# Ezek maradnak kötelezőek.
REQUIRED_HEADERS = {
    "Dátum",
    "Támadó",
    "Célország",
    "Régió",
    "Helyszín",
    "Szélesség",
    "Hosszúság",
    "Támadás típusa",
    "Célpont típusa",
    "Esemény rövid leírása",
    "Koordináta pontossága",
    "Forrás",
}


# Ezeket később hozzáadhatjuk az Excelhez,
# de hiányuk NEM akadályozza a feldolgozást.
OPTIONAL_HEADERS = {
    "Régió EN",
    "Helyszín EN",
    "Támadás típusa EN",
    "Célpont típusa EN",
    "Rövid leírás EN",
    "Részletes leírás HU",
    "Részletes leírás EN",
}


ATTACKER_MAP = {
    "ukrajna": "UKRAINE",
    "ukraine": "UKRAINE",
    "ua": "UKRAINE",

    "oroszország": "RUSSIA",
    "oroszorszag": "RUSSIA",
    "russia": "RUSSIA",
    "ru": "RUSSIA",
}


COUNTRY_MAP = {
    "ukrajna": "Ukraine",
    "ukraine": "Ukraine",

    "oroszország": "Russia",
    "oroszorszag": "Russia",
    "russia": "Russia",
}


# ============================================================
# SEGÉDFÜGGVÉNYEK
# ============================================================

def clean_text(value: Any) -> str:
    """
    Whitespace tisztítása és biztonságos szöveggé alakítás.
    """

    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value),
    ).strip()


def normalize_date(value: Any) -> str:
    """
    Dátum egységesítése YYYY-MM-DD formátumra.
    """

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    text = clean_text(value)

    formats = (
        "%Y-%m-%d",
        "%Y.%m.%d",
        "%Y.%m.%d.",
        "%d.%m.%Y",
        "%d.%m.%Y.",
        "%d/%m/%Y",
    )

    for fmt in formats:
        try:
            return datetime.strptime(
                text,
                fmt,
            ).date().isoformat()

        except ValueError:
            continue

    raise ValueError(
        f"Nem értelmezhető dátum: {value!r}"
    )


def normalize_attacker(value: Any) -> str:
    """
    Támadó fél normalizálása.
    """

    text = clean_text(value)

    if not text:
        return ""

    return ATTACKER_MAP.get(
        text.casefold(),
        text.upper(),
    )


def normalize_country(value: Any) -> str:
    """
    Célország egységesítése.
    """

    text = clean_text(value)

    if not text:
        return ""

    return COUNTRY_MAP.get(
        text.casefold(),
        text,
    )


def normalize_coordinate(
    value: Any,
    field_name: str,
) -> float:
    """
    Koordináta ellenőrzése és normalizálása.
    """

    if value is None or value == "":
        raise ValueError(
            f"Hiányzó koordináta: {field_name}"
        )

    try:
        number = float(value)

    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Hibás koordináta "
            f"({field_name}): {value!r}"
        ) from exc

    if not math.isfinite(number):
        raise ValueError(
            f"Nem véges koordináta "
            f"({field_name}): {value!r}"
        )

    if field_name == "latitude":
        if not -90 <= number <= 90:
            raise ValueError(
                f"Szélesség tartományon kívül: "
                f"{number}"
            )

    if field_name == "longitude":
        if not -180 <= number <= 180:
            raise ValueError(
                f"Hosszúság tartományon kívül: "
                f"{number}"
            )

    return round(number, 6)


def make_event_id(
    event_date: str,
    direction: str,
    location: str,
    latitude: float,
    longitude: float,
) -> str:
    """
    Stabil eseményazonosító.

    Nem Excel-sorszám alapján készül, ezért új sorok
    beszúrásakor a már meglévő események ID-je
    nem változik.
    """

    raw = (
        f"{event_date}|"
        f"{direction}|"
        f"{location.casefold()}|"
        f"{latitude:.6f}|"
        f"{longitude:.6f}"
    )

    digest = hashlib.sha1(
        raw.encode("utf-8")
    ).hexdigest()[:10].upper()

    date_code = event_date.replace(
        "-",
        "",
    )

    return (
        f"DEEP-{date_code}-"
        f"{direction}-{digest}"
    )


def make_duplicate_key(
    event_date: str,
    direction: str,
    location: str,
    latitude: float,
    longitude: float,
) -> tuple:
    """
    Duplikációk felismeréséhez használt kulcs.
    """

    return (
        event_date,
        direction,
        location.casefold(),
        round(latitude, 4),
        round(longitude, 4),
    )


def get_optional_value(
    row: tuple,
    column: dict[str, int],
    header_name: str,
) -> str:
    """
    Opcionális Excel-oszlop biztonságos olvasása.

    Ha az oszlop még nincs az Excelben,
    üres stringet ad vissza.
    """

    index = column.get(header_name)

    if index is None:
        return ""

    if index >= len(row):
        return ""

    return clean_text(
        row[index]
    )


# ============================================================
# EXCEL FELDOLGOZÁS
# ============================================================

def read_sheet(
    sheet: Any,
    sheet_name: str,
    direction: str,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
]:

    headers = [
        clean_text(cell.value)
        for cell in next(
            sheet.iter_rows(
                min_row=1,
                max_row=1,
            )
        )
    ]

    missing = sorted(
        REQUIRED_HEADERS - set(headers)
    )

    if missing:
        raise RuntimeError(
            f"A(z) '{sheet_name}' munkalapról "
            "hiányzó kötelező oszlopok: "
            + ", ".join(missing)
        )

    column = {
        name: index
        for index, name in enumerate(headers)
        if name
    }

    events: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    duplicate_keys: set[tuple] = set()

    for row_number, row in enumerate(
        sheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):

        # Teljesen üres sor kihagyása.
        if not any(
            value not in (None, "")
            for value in row
        ):
            continue

        try:

            # ====================================================
            # ALAP ADATOK
            # ====================================================

            event_date = normalize_date(
                row[column["Dátum"]]
            )

            attacker = normalize_attacker(
                row[column["Támadó"]]
            )

            target_country = normalize_country(
                row[column["Célország"]]
            )

            region_hu = clean_text(
                row[column["Régió"]]
            )

            location_hu = clean_text(
                row[column["Helyszín"]]
            )

            latitude = normalize_coordinate(
                row[column["Szélesség"]],
                "latitude",
            )

            longitude = normalize_coordinate(
                row[column["Hosszúság"]],
                "longitude",
            )

            strike_type_hu = clean_text(
                row[
                    column[
                        "Támadás típusa"
                    ]
                ]
            )

            target_type_hu = clean_text(
                row[
                    column[
                        "Célpont típusa"
                    ]
                ]
            )

            description_short_hu = clean_text(
                row[
                    column[
                        "Esemény rövid leírása"
                    ]
                ]
            )

            coordinate_accuracy = clean_text(
                row[
                    column[
                        "Koordináta pontossága"
                    ]
                ]
            )

            source_url = clean_text(
                row[
                    column[
                        "Forrás"
                    ]
                ]
            )


            # ====================================================
            # OPCIONÁLIS ANGOL ADATOK
            # ====================================================

            region_en = get_optional_value(
                row,
                column,
                "Régió EN",
            )

            location_en = get_optional_value(
                row,
                column,
                "Helyszín EN",
            )

            strike_type_en = get_optional_value(
                row,
                column,
                "Támadás típusa EN",
            )

            target_type_en = get_optional_value(
                row,
                column,
                "Célpont típusa EN",
            )

            description_short_en = (
                get_optional_value(
                    row,
                    column,
                    "Rövid leírás EN",
                )
            )


            # ====================================================
            # RÉSZLETES LEÍRÁS
            # ====================================================

            description_long_hu = (
                get_optional_value(
                    row,
                    column,
                    "Részletes leírás HU",
                )
            )

            description_long_en = (
                get_optional_value(
                    row,
                    column,
                    "Részletes leírás EN",
                )
            )

            # Ha még nincs külön részletes magyar szöveg,
            # legalább a rövid leírás megjelenik lenyitáskor.
            if not description_long_hu:
                description_long_hu = (
                    description_short_hu
                )

            # Angol részletes szöveg csak akkor öröklődik
            # a rövid angolból, ha az létezik.
            if (
                not description_long_en
                and description_short_en
            ):
                description_long_en = (
                    description_short_en
                )


            # ====================================================
            # KÖTELEZŐ MEZŐK ELLENŐRZÉSE
            # ====================================================

            required_fields = {
                "Támadó": attacker,
                "Célország": target_country,
                "Régió": region_hu,
                "Helyszín": location_hu,
                "Támadás típusa": (
                    strike_type_hu
                ),
                "Célpont típusa": (
                    target_type_hu
                ),
                "Esemény rövid leírása": (
                    description_short_hu
                ),
                "Forrás": source_url,
            }

            empty_fields = [
                name
                for name, value
                in required_fields.items()
                if not value
            ]

            if empty_fields:
                raise ValueError(
                    "Hiányzó kötelező mező(k): "
                    + ", ".join(empty_fields)
                )


            # ====================================================
            # IRÁNY ELLENŐRZÉSE
            # ====================================================

            expected_attacker = (
                "UKRAINE"
                if direction == "UA_RU"
                else "RUSSIA"
            )

            expected_country = (
                "Russia"
                if direction == "UA_RU"
                else "Ukraine"
            )

            if attacker != expected_attacker:
                warnings.append(
                    {
                        "sheet": sheet_name,
                        "row": row_number,
                        "warning": (
                            "A támadó fél nem egyezik "
                            "a munkalap irányával. "
                            f"Várt: {expected_attacker}, "
                            f"kapott: {attacker}"
                        ),
                    }
                )

            if target_country != expected_country:
                warnings.append(
                    {
                        "sheet": sheet_name,
                        "row": row_number,
                        "warning": (
                            "A célország nem egyezik "
                            "a munkalap irányával. "
                            f"Várt: {expected_country}, "
                            f"kapott: {target_country}"
                        ),
                    }
                )


            # ====================================================
            # FORRÁS ELLENŐRZÉS
            # ====================================================

            if not source_url.startswith(
                (
                    "http://",
                    "https://",
                )
            ):
                warnings.append(
                    {
                        "sheet": sheet_name,
                        "row": row_number,
                        "warning": (
                            "A Forrás mező nem "
                            "http:// vagy https:// "
                            "címmel kezdődik."
                        ),
                    }
                )


            # ====================================================
            # DUPLIKÁCIÓ
            # ====================================================

            duplicate_key = (
                make_duplicate_key(
                    event_date,
                    direction,
                    location_hu,
                    latitude,
                    longitude,
                )
            )

            if duplicate_key in duplicate_keys:
                warnings.append(
                    {
                        "sheet": sheet_name,
                        "row": row_number,
                        "warning": (
                            "Lehetséges duplikált "
                            "esemény ugyanazon "
                            "dátummal, helyszínnel "
                            "és koordinátával."
                        ),
                    }
                )

            duplicate_keys.add(
                duplicate_key
            )


            # ====================================================
            # EVENT ID
            # ====================================================

            event_id = make_event_id(
                event_date,
                direction,
                location_hu,
                latitude,
                longitude,
            )


            # ====================================================
            # JSON REKORD
            # ====================================================

            event = {

                # ------------------------------------------------
                # AZONOSÍTÁS
                # ------------------------------------------------

                "event_id": event_id,

                "date": event_date,

                "direction": direction,

                "attacker": attacker,

                "target_country": target_country,


                # ------------------------------------------------
                # VISSZAFELÉ KOMPATIBILIS MEZŐK
                #
                # Ezeket a jelenlegi térképi rendszer már
                # használja, ezért NEM töröljük őket.
                # ------------------------------------------------

                "region": region_hu,

                "location": location_hu,

                "strike_type": strike_type_hu,

                "target_type": target_type_hu,

                "description": (
                    description_short_hu
                ),


                # ------------------------------------------------
                # KÉTNYELVŰ HELYSZÍN
                # ------------------------------------------------

                "region_hu": region_hu,

                "region_en": region_en,

                "location_hu": location_hu,

                "location_en": location_en,


                # ------------------------------------------------
                # KÉTNYELVŰ TÁMADÁSTÍPUS
                # ------------------------------------------------

                "strike_type_hu": (
                    strike_type_hu
                ),

                "strike_type_en": (
                    strike_type_en
                ),


                # ------------------------------------------------
                # KÉTNYELVŰ CÉLPONTTÍPUS
                # ------------------------------------------------

                "target_type_hu": (
                    target_type_hu
                ),

                "target_type_en": (
                    target_type_en
                ),


                # ------------------------------------------------
                # RÖVID LEÍRÁS
                #
                # Ez kerül a kompakt térképi kártyára.
                # ------------------------------------------------

                "description_short_hu": (
                    description_short_hu
                ),

                "description_short_en": (
                    description_short_en
                ),


                # ------------------------------------------------
                # RÉSZLETES LEÍRÁS
                #
                # Ez jelenik meg a ▼ lenyitás után.
                # ------------------------------------------------

                "description_long_hu": (
                    description_long_hu
                ),

                "description_long_en": (
                    description_long_en
                ),


                # ------------------------------------------------
                # KOORDINÁTÁK
                # ------------------------------------------------

                "latitude": latitude,

                "longitude": longitude,

                "coordinate_accuracy": (
                    coordinate_accuracy
                ),


                # ------------------------------------------------
                # FORRÁS
                # ------------------------------------------------

                "source_url": source_url,


                # ------------------------------------------------
                # DEBUG / EREDET
                # ------------------------------------------------

                "source_sheet": sheet_name,

                "source_row": row_number,
            }

            events.append(event)


        except Exception as exc:

            warnings.append(
                {
                    "sheet": sheet_name,
                    "row": row_number,
                    "error": str(exc),
                }
            )

    return events, warnings


def read_events(
    input_path: Path,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
]:

    workbook = load_workbook(
        input_path,
        data_only=True,
        read_only=True,
    )

    events: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    for (
        sheet_name,
        direction,
    ) in SHEETS.items():

        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(
                "Az Excel nem tartalmazza "
                "a szükséges "
                f"'{sheet_name}' munkalapot."
            )

        sheet = workbook[
            sheet_name
        ]

        (
            sheet_events,
            sheet_warnings,
        ) = read_sheet(
            sheet,
            sheet_name,
            direction,
        )

        events.extend(
            sheet_events
        )

        warnings.extend(
            sheet_warnings
        )

    events.sort(
        key=lambda item: (
            item["date"],
            item["direction"],
            item["region"],
            item["location"],
            item["event_id"],
        )
    )

    return events, warnings


# ============================================================
# ÖSSZESÍTÉS
# ============================================================

def build_summary(
    events: list[dict[str, Any]],
) -> dict[str, Any]:

    by_direction = Counter(
        event["direction"]
        for event in events
    )

    by_attacker = Counter(
        event["attacker"]
        for event in events
    )

    by_target_country = Counter(
        event["target_country"]
        for event in events
    )

    by_region = Counter(
        event["region"]
        for event in events
    )

    by_strike_type = Counter(
        event["strike_type"]
        for event in events
    )

    by_target_type = Counter(
        event["target_type"]
        for event in events
    )

    by_date = Counter(
        event["date"]
        for event in events
    )

    dates = [
        event["date"]
        for event in events
    ]


    # ========================================================
    # KÉTNYELVŰ ADATOK LEFEDETTSÉGE
    # ========================================================

    english_ready = sum(
        1
        for event in events
        if (
            event.get("location_en")
            and event.get("strike_type_en")
            and event.get("target_type_en")
            and event.get(
                "description_short_en"
            )
        )
    )

    detailed_hu_ready = sum(
        1
        for event in events
        if event.get(
            "description_long_hu"
        )
    )

    detailed_en_ready = sum(
        1
        for event in events
        if event.get(
            "description_long_en"
        )
    )


    return {

        "event_count": len(events),

        "date_start": (
            min(dates)
            if dates
            else None
        ),

        "date_end": (
            max(dates)
            if dates
            else None
        ),

        "directions": dict(
            sorted(
                by_direction.items()
            )
        ),

        "attackers": dict(
            sorted(
                by_attacker.items()
            )
        ),

        "target_countries": dict(
            sorted(
                by_target_country.items()
            )
        ),

        "regions": dict(
            sorted(
                by_region.items(),
                key=lambda item: (
                    -item[1],
                    item[0],
                ),
            )
        ),

        "strike_types": dict(
            sorted(
                by_strike_type.items(),
                key=lambda item: (
                    -item[1],
                    item[0],
                ),
            )
        ),

        "target_types": dict(
            sorted(
                by_target_type.items(),
                key=lambda item: (
                    -item[1],
                    item[0],
                ),
            )
        ),

        "daily_counts": dict(
            sorted(
                by_date.items()
            )
        ),


        # ----------------------------------------------------
        # Nyelvi lefedettség
        # ----------------------------------------------------

        "language_coverage": {

            "english_complete_events": (
                english_ready
            ),

            "english_complete_percent": (
                round(
                    (
                        english_ready
                        / len(events)
                        * 100
                    ),
                    1,
                )
                if events
                else 0
            ),

            "detailed_hu_events": (
                detailed_hu_ready
            ),

            "detailed_en_events": (
                detailed_en_ready
            ),
        },
    }


# ============================================================
# JSON ÍRÁS
# ============================================================

def write_json(
    path: Path,
    payload: Any,
) -> None:

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    path.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


# ============================================================
# MAIN
# ============================================================

def main() -> int:

    parser = argparse.ArgumentParser(
        description=(
            "Ukrajna–Oroszország "
            "mélységi csapásadatok "
            "Excel → JSON feldolgozó"
        )
    )

    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT,
    )

    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
    )

    parser.add_argument(
        "--summary",
        default=DEFAULT_SUMMARY,
    )

    parser.add_argument(
        "--validation",
        default=DEFAULT_VALIDATION,
    )

    args = parser.parse_args()

    input_path = Path(
        args.input
    )

    if not input_path.exists():

        print(
            "HIBA: Nem található "
            "a bemeneti Excel: "
            f"{input_path}",
            file=sys.stderr,
        )

        return 1


    try:

        events, warnings = (
            read_events(
                input_path
            )
        )

    except Exception as exc:

        print(
            f"HIBA: {exc}",
            file=sys.stderr,
        )

        return 1


    errors = [
        item
        for item in warnings
        if "error" in item
    ]

    warning_items = [
        item
        for item in warnings
        if "warning" in item
    ]


    generated_at = (
        datetime.now(
            timezone.utc
        )
        .replace(
            microsecond=0
        )
        .isoformat()
        .replace(
            "+00:00",
            "Z",
        )
    )


    summary = build_summary(
        events
    )


    dataset = {

        "generated_at": (
            generated_at
        ),

        "dataset": (
            "ukraine_russia_"
            "deep_strike_history"
        ),

        "source_file": (
            input_path.name
        ),

        "source_path": (
            str(input_path)
        ),

        "update_mode": "manual",

        "schema_version": "2.0",

        "multilingual": True,

        "languages": [
            "hu",
            "en",
        ],

        "summary": summary,

        "events": events,
    }


    validation = {

        "generated_at": (
            generated_at
        ),

        "source_file": (
            input_path.name
        ),

        "source_path": (
            str(input_path)
        ),

        "schema_version": "2.0",

        "valid_event_count": (
            len(events)
        ),

        "error_count": (
            len(errors)
        ),

        "warning_count": (
            len(warning_items)
        ),

        "language_coverage": (
            summary.get(
                "language_coverage",
                {},
            )
        ),

        "items": warnings,
    }


    write_json(
        Path(
            args.output
        ),
        dataset,
    )


    write_json(
        Path(
            args.summary
        ),
        {
            "generated_at": (
                generated_at
            ),

            "dataset": (
                "ukraine_russia_"
                "deep_strike_history"
            ),

            "schema_version": "2.0",

            **summary,
        },
    )


    write_json(
        Path(
            args.validation
        ),
        validation,
    )


    print(
        "Feldolgozott események: "
        f"{len(events)}"
    )

    print(
        "Ukrajna → Oroszország: "
        f"{summary['directions'].get('UA_RU', 0)}"
    )

    print(
        "Oroszország → Ukrajna: "
        f"{summary['directions'].get('RU_UA', 0)}"
    )

    language_coverage = (
        summary.get(
            "language_coverage",
            {},
        )
    )

    print(
        "Teljes angol rekordok: "
        f"{language_coverage.get('english_complete_events', 0)}"
        "/"
        f"{len(events)}"
    )

    print(
        "Részletes HU leírások: "
        f"{language_coverage.get('detailed_hu_events', 0)}"
    )

    print(
        "Részletes EN leírások: "
        f"{language_coverage.get('detailed_en_events', 0)}"
    )

    print(
        f"Hibák: {len(errors)} | "
        f"Figyelmeztetések: "
        f"{len(warning_items)}"
    )

    print(
        f"Kimenet: {args.output}"
    )

    print(
        f"Összesítés: {args.summary}"
    )

    print(
        f"Validáció: "
        f"{args.validation}"
    )


    if errors:

        print(
            "HIBA: Egy vagy több "
            "Excel-sor nem volt "
            "feldolgozható.",
            file=sys.stderr,
        )

        return 2


    return 0


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
