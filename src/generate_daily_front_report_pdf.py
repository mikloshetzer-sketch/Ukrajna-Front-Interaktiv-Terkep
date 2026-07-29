#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
TÖRÉSVONALAK INTELLIGENCE HUB
Ukrajnai Fronthelyzet - napi stratégiai hírszerzési PDF-jelentés

Fájl helye a repóban:
    src/generate_daily_front_report_pdf.py

Kimenetek:
    docs/reports/latest-hu.pdf
    docs/reports/latest-en.pdf
    docs/reports/archive/YYYY-MM-DD-hu.pdf
    docs/reports/archive/YYYY-MM-DD-en.pdf
    docs/reports/reports_index.json

A generátor minden futáskor újraépíti az adott napi archív PDF-et.
A heti vagy időszakos adatokat nem kezeli napi adatként. Minden blokk
adatfrissességi státuszt és lefedett időszakot kap.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import shutil
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    HRFlowable,
    KeepTogether,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


VERSION = "1.0.0"
ROOT = Path(__file__).resolve().parents[1]

DATA_DIR = ROOT / "data"
DOCS_DIR = ROOT / "docs"
REPORT_DIR = DOCS_DIR / "reports"
ARCHIVE_DIR = REPORT_DIR / "archive"

INPUTS = {
    "long_term": DATA_DIR / "dashboard_long_term_summary.json",
    "osint": DATA_DIR / "osint_feed.json",
    "telegram": DATA_DIR / "telegram_feed.json",
    "front_summary": DATA_DIR / "suriyak_front_summary.json",
    "territorial_daily": DATA_DIR / "territorial_delta.geojson",
    "territorial_30": DATA_DIR / "territorial_delta_30days.geojson",
    "territorial_windows": DATA_DIR / "territorial_delta_windows.geojson",
    "territorial_history": DATA_DIR / "territorial_history_daily.json",
    "firms_1": DATA_DIR / "firms_1.json",
    "firms_3": DATA_DIR / "firms_3.json",
    "firms_10": DATA_DIR / "firms_10.json",
    "firms_30": DATA_DIR / "firms_30.json",
    "units": DATA_DIR / "unit_feed.json",
    "deep_strikes": DATA_DIR / "manual" / "ukrajna_oroszorszag_melysegi_csapasok.xlsx",
}

NAVY = colors.HexColor("#0B2945")
BLUE = colors.HexColor("#1D5A88")
MID_BLUE = colors.HexColor("#2E6F9E")
LIGHT_BLUE = colors.HexColor("#EAF2F8")
PALE_BLUE = colors.HexColor("#F4F8FB")
DARK = colors.HexColor("#12263A")
TEXT = colors.HexColor("#1E2B37")
MUTED = colors.HexColor("#5E7182")
BORDER = colors.HexColor("#C9D7E3")
GREEN = colors.HexColor("#2F7D57")
AMBER = colors.HexColor("#A86B00")
RED = colors.HexColor("#A64040")
GREY = colors.HexColor("#E9EEF2")
WHITE = colors.white

HU_MONTHS = (
    "január", "február", "március", "április", "május", "június",
    "július", "augusztus", "szeptember", "október", "november", "december",
)


@dataclass
class Freshness:
    frequency: str
    updated_at: datetime | None
    period_start: date | None
    period_end: date | None
    age_days: int | None
    status: str
    include_in_daily: bool
    note: str


@dataclass
class ReportPaths:
    archive: Path
    latest: Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--lang", choices=("hu", "en", "all"), default="all")
    parser.add_argument("--report-date", help="YYYY-MM-DD; alapértelmezés: automatikus")
    parser.add_argument("--output-dir", default=str(REPORT_DIR))
    parser.add_argument("--force", action="store_true", help="Kompatibilitási kapcsoló; a fájl mindig újraépül.")
    return parser.parse_args()


def read_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return {} if default is None else default
    try:
        with path.open("r", encoding="utf-8-sig") as handle:
            return json.load(handle)
    except Exception as exc:
        print(f"WARNING: Nem olvasható JSON: {path}: {exc}", file=sys.stderr)
        return {} if default is None else default


def nested(obj: Any, *keys: str, default: Any = None) -> Any:
    current = obj
    for key in keys:
        if not isinstance(current, Mapping) or key not in current:
            return default
        current = current[key]
    return current


def first_value(obj: Any, paths: Sequence[Sequence[str]], default: Any = None) -> Any:
    for path in paths:
        value = nested(obj, *path, default=None)
        if value is not None and value != "":
            return value
    return default


def parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime(value.year, value.month, value.day, tzinfo=timezone.utc)
    elif isinstance(value, (int, float)):
        try:
            dt = datetime.fromtimestamp(value, tz=timezone.utc)
        except Exception:
            return None
    else:
        text = str(value).strip()
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        for parser in (
            lambda s: datetime.fromisoformat(s),
            lambda s: datetime.strptime(s[:10], "%Y-%m-%d"),
            lambda s: datetime.strptime(s[:10], "%Y.%m.%d"),
        ):
            try:
                dt = parser(text)
                break
            except Exception:
                dt = None
        if dt is None:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def parse_date(value: Any) -> date | None:
    dt = parse_datetime(value)
    return dt.date() if dt else None


def to_float(value: Any, default: float = 0.0) -> float:
    if isinstance(value, bool):
        return default
    try:
        if value is None or value == "":
            return default
        return float(str(value).replace(",", "."))
    except Exception:
        return default


def to_int(value: Any, default: int = 0) -> int:
    try:
        return int(round(to_float(value, float(default))))
    except Exception:
        return default


def esc(value: Any) -> str:
    text = "" if value is None else str(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def human_date(day: date, lang: str) -> str:
    if lang == "hu":
        return f"{day.year}. {HU_MONTHS[day.month - 1]} {day.day}."
    return day.strftime("%B %d, %Y")


def num(value: float, digits: int = 2, signed: bool = False) -> str:
    if not math.isfinite(value):
        return "-"
    prefix = "+" if signed and value > 0 else ""
    return f"{prefix}{value:.{digits}f}"


def find_dates_recursive(obj: Any) -> list[datetime]:
    found: list[datetime] = []
    date_keys = {
        "generated_at", "updated_at", "last_updated", "timestamp", "date",
        "as_of", "report_date", "period_end", "end_date", "created_at",
    }

    def walk(node: Any, parent_key: str = "") -> None:
        if isinstance(node, Mapping):
            for key, value in node.items():
                key_lower = str(key).lower()
                if key_lower in date_keys or "date" in key_lower or "time" in key_lower:
                    dt = parse_datetime(value)
                    if dt:
                        found.append(dt)
                if isinstance(value, (Mapping, list)):
                    walk(value, key_lower)
        elif isinstance(node, list):
            for item in node[:5000]:
                walk(item, parent_key)

    walk(obj)
    return found


def latest_date_in_data(obj: Any, fallback_path: Path | None = None) -> datetime | None:
    dates = find_dates_recursive(obj)
    if dates:
        return max(dates)
    if fallback_path and fallback_path.exists():
        return datetime.fromtimestamp(fallback_path.stat().st_mtime, tz=timezone.utc)
    return None


def infer_report_date(datasets: Mapping[str, Any], requested: str | None) -> date:
    if requested:
        return datetime.strptime(requested, "%Y-%m-%d").date()

    candidates: list[datetime] = []
    for key in ("long_term", "territorial_daily", "osint", "telegram", "firms_1"):
        obj = datasets.get(key)
        path = INPUTS.get(key)
        dt = latest_date_in_data(obj, path)
        if dt:
            candidates.append(dt)
    return max(candidates).date() if candidates else datetime.now(timezone.utc).date()


def feature_list(geojson: Any) -> list[dict[str, Any]]:
    if isinstance(geojson, Mapping):
        features = geojson.get("features")
        if isinstance(features, list):
            return [x for x in features if isinstance(x, Mapping)]
    return []


def extract_items(obj: Any) -> list[dict[str, Any]]:
    if isinstance(obj, list):
        return [x for x in obj if isinstance(x, Mapping)]
    if not isinstance(obj, Mapping):
        return []
    for key in ("items", "events", "articles", "records", "data", "features", "results"):
        value = obj.get(key)
        if isinstance(value, list):
            return [x for x in value if isinstance(x, Mapping)]
    return []


def extract_period(obj: Any) -> tuple[date | None, date | None]:
    start = first_value(
        obj,
        (
            ("period_start",), ("start_date",), ("from",), ("coverage_start",),
            ("metadata", "period_start"), ("metadata", "start_date"),
        ),
    )
    end = first_value(
        obj,
        (
            ("period_end",), ("end_date",), ("to",), ("coverage_end",),
            ("metadata", "period_end"), ("metadata", "end_date"),
        ),
    )
    start_date = parse_date(start)
    end_date = parse_date(end)

    if start_date is None or end_date is None:
        dates = [parse_date(x.get("date")) for x in extract_items(obj)]
        dates = [x for x in dates if x]
        if dates:
            start_date = start_date or min(dates)
            end_date = end_date or max(dates)
    return start_date, end_date


def freshness_for(
    obj: Any,
    path: Path,
    report_date: date,
    frequency: str,
    max_current_days: int,
    max_usable_days: int,
    lang: str,
) -> Freshness:
    updated = latest_date_in_data(obj, path)
    start, end = extract_period(obj)
    reference = end or (updated.date() if updated else None)
    age = (report_date - reference).days if reference else None

    if age is None:
        status = "missing"
        include = False
    elif age <= max_current_days:
        status = "current"
        include = True
    elif age <= max_usable_days:
        status = "older"
        include = True
    else:
        status = "stale"
        include = False

    if lang == "hu":
        notes = {
            "current": "A legutóbbi elérhető adatállomány.",
            "older": "Korábbi adatállomány; csak háttérjelzésként értelmezendő.",
            "stale": "Elavult adat; a napi integrált értékelés nem használja.",
            "missing": "Nem áll rendelkezésre értelmezhető frissítési dátum.",
        }
    else:
        notes = {
            "current": "Latest available dataset.",
            "older": "Older dataset; used only as contextual evidence.",
            "stale": "Stale dataset; excluded from the integrated daily assessment.",
            "missing": "No interpretable update date is available.",
        }

    return Freshness(
        frequency=frequency,
        updated_at=updated,
        period_start=start,
        period_end=end,
        age_days=age,
        status=status,
        include_in_daily=include,
        note=notes[status],
    )


def deep_strike_rows(path: Path) -> tuple[list[dict[str, Any]], datetime | None]:
    if not path.exists() or load_workbook is None:
        return [], None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc)

        headers = [str(x).strip() if x is not None else f"col_{idx}" for idx, x in enumerate(rows[0])]
        records: list[dict[str, Any]] = []
        for values in rows[1:]:
            if not any(v is not None and str(v).strip() for v in values):
                continue
            records.append({headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))})
        return records, datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc)
    except Exception as exc:
        print(f"WARNING: Nem olvasható Excel: {path}: {exc}", file=sys.stderr)
        return [], None


def normalize_deep_strikes(path: Path) -> dict[str, Any]:
    rows, mtime = deep_strike_rows(path)
    result: dict[str, Any] = {"items": rows}
    if mtime:
        result["updated_at"] = mtime.isoformat()

    dates: list[date] = []
    for row in rows:
        for key, value in row.items():
            if any(token in str(key).lower() for token in ("dátum", "datum", "date", "időpont")):
                parsed = parse_date(value)
                if parsed:
                    dates.append(parsed)
                    break
    if dates:
        result["period_start"] = min(dates).isoformat()
        result["period_end"] = max(dates).isoformat()
    return result


def territorial_metrics(obj: Any) -> dict[str, Any]:
    features = feature_list(obj)
    ru_gain = 0.0
    ua_gain = 0.0
    rows: list[dict[str, Any]] = []

    for feature in features:
        props = dict(feature.get("properties") or {})
        area = to_float(first_value(
            props,
            (
                ("area_km2",), ("area_sq_km",), ("km2",), ("area",),
                ("delta_km2",), ("territorial_change_km2",),
            ),
        ))
        blob = " ".join(str(v) for v in props.values()).lower()
        side = str(first_value(props, (("side",), ("actor",), ("type",), ("control",)), default="")).lower()

        is_ua = any(x in side or x in blob for x in ("ukrain", "ua gain", "ua recapture", "ukrán", "ukrainian"))
        is_ru = any(x in side or x in blob for x in ("russian", "ru gain", "orosz", "russia"))

        if is_ua and not is_ru:
            ua_gain += abs(area)
            actor = "UA"
        else:
            ru_gain += abs(area)
            actor = "RU"

        place = first_value(
            props,
            (
                ("nearest_place",), ("nearestPlace",), ("location",), ("place",),
                ("name",), ("sector",), ("sectorName",),
            ),
            default="-",
        )
        sector = first_value(
            props,
            (("sector",), ("sectorName",), ("sector_name",), ("front_sector",)),
            default="-",
        )
        rows.append({"actor": actor, "area": abs(area), "place": place, "sector": sector, "properties": props})

    summary = obj if isinstance(obj, Mapping) else {}
    ru_gain = ru_gain or to_float(first_value(
        summary,
        (
            ("ru_gain_km2",), ("russian_gain_km2",), ("summary", "ru_gain_km2"),
            ("totals", "russian_gain_km2"), ("statistics", "ru_gain_km2"),
        ),
    ))
    ua_gain = ua_gain or to_float(first_value(
        summary,
        (
            ("ua_gain_km2",), ("ukrainian_gain_km2",), ("summary", "ua_gain_km2"),
            ("totals", "ukrainian_gain_km2"), ("statistics", "ua_gain_km2"),
        ),
    ))
    return {
        "ru_gain": ru_gain,
        "ua_gain": ua_gain,
        "net_ru": ru_gain - ua_gain,
        "rows": sorted(rows, key=lambda x: x["area"], reverse=True),
        "feature_count": len(features),
    }


def firms_points(obj: Any) -> list[dict[str, Any]]:
    items = extract_items(obj)
    result: list[dict[str, Any]] = []
    for item in items:
        props = item.get("properties") if isinstance(item.get("properties"), Mapping) else item
        lat = first_value(props, (("latitude",), ("lat",), ("y",)))
        lon = first_value(props, (("longitude",), ("lon",), ("lng",), ("x",)))
        result.append({
            "lat": to_float(lat, math.nan),
            "lon": to_float(lon, math.nan),
            "date": first_value(props, (("date",), ("acq_date",), ("timestamp",))),
            "brightness": to_float(first_value(props, (("brightness",), ("bright_ti4",), ("frp",))), 0.0),
            "place": first_value(props, (("nearestPlace",), ("place",), ("location",)), default="-"),
            "sector": first_value(props, (("sectorName",), ("sector",), ("front_sector",)), default="-"),
        })
    return result


def osint_metrics(*datasets: Any) -> dict[str, Any]:
    items: list[dict[str, Any]] = []
    for dataset in datasets:
        items.extend(extract_items(dataset))

    unique: dict[str, dict[str, Any]] = {}
    for item in items:
        key = str(item.get("url") or item.get("title") or json.dumps(item, sort_keys=True, default=str))
        unique[key] = item
    items = list(unique.values())

    sector_counts: Counter[str] = Counter()
    category_counts: Counter[str] = Counter()
    importance_total = 0.0
    for item in items:
        sector = first_value(item, (("sectorName",), ("sectorShortName",), ("sector",)), default="Általános")
        category = first_value(item, (("category",), ("type",), ("event_type",)), default="egyéb")
        sector_counts[str(sector)] += 1
        category_counts[str(category)] += 1
        importance_total += to_float(item.get("importance"), 0.0)

    items.sort(
        key=lambda x: (
            to_float(x.get("importance"), 0.0),
            str(x.get("date") or x.get("published_at") or ""),
        ),
        reverse=True,
    )
    return {
        "items": items,
        "count": len(items),
        "sector_counts": sector_counts,
        "category_counts": category_counts,
        "avg_importance": importance_total / len(items) if items else 0.0,
    }


def unit_metrics(obj: Any) -> dict[str, Any]:
    items = extract_items(obj)
    sides = Counter()
    sectors = Counter()
    for item in items:
        side = str(first_value(item, (("side",), ("actor",), ("country",)), default="Ismeretlen"))
        sector = str(first_value(item, (("sector",), ("sectorName",), ("area",)), default="Ismeretlen"))
        sides[side] += 1
        sectors[sector] += 1
    return {"items": items, "count": len(items), "sides": sides, "sectors": sectors}


def long_term_metrics(obj: Any) -> dict[str, Any]:
    return {
        "threat_sector": first_value(
            obj,
            (
                ("highest_threat_sector",), ("summary", "highest_threat_sector"),
                ("assessment", "highest_threat_sector"), ("front", "highest_threat_sector"),
            ),
            default="-",
        ),
        "threat_level": first_value(
            obj,
            (
                ("highest_threat_level",), ("summary", "highest_threat_level"),
                ("assessment", "highest_threat_level"), ("front", "highest_threat_level"),
            ),
            default="-",
        ),
        "active_sector": first_value(
            obj,
            (
                ("most_active_sector",), ("summary", "most_active_sector"),
                ("assessment", "most_active_sector"), ("front", "most_active_sector"),
            ),
            default="-",
        ),
        "conflict_index": to_float(first_value(
            obj,
            (
                ("conflict_index",), ("summary", "conflict_index"),
                ("assessment", "conflict_index"), ("daily", "conflict_index"),
            ),
        )),
        "ru_control": to_float(first_value(
            obj,
            (
                ("russian_control_km2",), ("summary", "russian_control_km2"),
                ("territory", "russian_control_km2"), ("current", "russian_control_km2"),
            ),
        )),
        "daily_events": to_int(first_value(
            obj,
            (
                ("daily_events",), ("summary", "daily_events"),
                ("events", "daily_count"), ("daily", "events"),
            ),
        )),
    }


def load_datasets() -> dict[str, Any]:
    datasets = {key: read_json(path) for key, path in INPUTS.items() if path.suffix.lower() != ".xlsx"}
    datasets["deep_strikes"] = normalize_deep_strikes(INPUTS["deep_strikes"])
    return datasets


def i18n(lang: str) -> dict[str, str]:
    if lang == "hu":
        return {
            "title": "Ukrajnai Fronthelyzet - Napi Stratégiai Hírszerzési Jelentés",
            "short_title": "Ukrajnai Fronthelyzet",
            "subtitle": "Nyílt forrású információk integrált, számszerű és szöveges értékelése",
            "about": "A jelentés célja",
            "about_text": (
                "A dokumentum a frontvonal, a területi változások, a harcintenzitás, a FIRMS-hőpontok, "
                "az OSINT-források, az egységadatok és a mélységi csapások adatait egyetlen elemzési "
                "keretbe rendezi. A napi, gördülő és heti adatokat elkülönítve kezeli, így egy régebbi "
                "vagy ritkábban frissülő adat nem jelenik meg automatikusan napi fejleményként."
            ),
            "input_models": "Adatforrások és frissesség",
            "glance": "A jelentés egy pillantásra",
            "executive": "Vezetői összefoglaló",
            "territory": "Területi helyzet",
            "activity": "Frontaktivitás és harcintenzitás",
            "firms": "FIRMS-hőpontok és rövid távú aktivitás",
            "deep": "Mélységi csapások - legutóbbi elérhető heti összkép",
            "units": "Egységek és erőcsoportosítás",
            "osint": "OSINT-helyzetkép és napi mozgatórugók",
            "integrated": "Integrált stratégiai értékelés",
            "outlook": "24-72 órás kilátások",
            "method": "Módszertan és korlátok",
            "disclaimer": "Jogi és módszertani nyilatkozat",
            "generated": "Jelentés dátuma",
            "frequency": "Frissítési gyakoriság",
            "updated": "Utolsó frissítés",
            "coverage": "Lefedett időszak",
            "status": "Státusz",
            "daily": "napi",
            "rolling": "gördülő",
            "weekly": "heti",
            "periodic": "időszakos",
        }
    return {
        "title": "Ukraine Frontline - Daily Strategic Intelligence Report",
        "short_title": "Ukraine Frontline",
        "subtitle": "Integrated quantitative and narrative assessment of open-source information",
        "about": "Purpose of this Report",
        "about_text": (
            "The report integrates frontline data, territorial changes, combat intensity, FIRMS hotspots, "
            "OSINT reporting, unit observations and deep-strike records. Daily, rolling and weekly datasets "
            "are treated separately, preventing older or less frequently updated information from being "
            "misrepresented as a same-day development."
        ),
        "input_models": "Data Sources and Freshness",
        "glance": "Report at a Glance",
        "executive": "Executive Summary",
        "territory": "Territorial Situation",
        "activity": "Frontline Activity and Combat Intensity",
        "firms": "FIRMS Hotspots and Short-Term Activity",
        "deep": "Deep Strikes - Latest Available Weekly Picture",
        "units": "Units and Force Posture",
        "osint": "OSINT Situation and Daily Drivers",
        "integrated": "Integrated Strategic Assessment",
        "outlook": "24-72 Hour Outlook",
        "method": "Methodology and Limitations",
        "disclaimer": "Legal and Methodological Disclaimer",
        "generated": "Report date",
        "frequency": "Update frequency",
        "updated": "Last update",
        "coverage": "Coverage period",
        "status": "Status",
        "daily": "daily",
        "rolling": "rolling",
        "weekly": "weekly",
        "periodic": "periodic",
    }


def make_styles() -> dict[str, ParagraphStyle]:
    sample = getSampleStyleSheet()
    return {
        "cover_brand": ParagraphStyle(
            "cover_brand", parent=sample["Normal"], fontName="Helvetica-Bold",
            fontSize=15, leading=18, textColor=WHITE, alignment=TA_CENTER,
        ),
        "cover_title": ParagraphStyle(
            "cover_title", parent=sample["Title"], fontName="Helvetica-Bold",
            fontSize=26, leading=31, textColor=WHITE, alignment=TA_CENTER,
            spaceAfter=7 * mm,
        ),
        "cover_subtitle": ParagraphStyle(
            "cover_subtitle", parent=sample["Normal"], fontName="Helvetica",
            fontSize=11.5, leading=17, textColor=colors.HexColor("#DCEAF4"),
            alignment=TA_CENTER,
        ),
        "h1": ParagraphStyle(
            "h1", parent=sample["Heading1"], fontName="Helvetica-Bold",
            fontSize=20, leading=24, textColor=NAVY, spaceBefore=5 * mm,
            spaceAfter=4 * mm, keepWithNext=True,
        ),
        "h2": ParagraphStyle(
            "h2", parent=sample["Heading2"], fontName="Helvetica-Bold",
            fontSize=14, leading=18, textColor=BLUE, spaceBefore=4 * mm,
            spaceAfter=2.5 * mm, keepWithNext=True,
        ),
        "h3": ParagraphStyle(
            "h3", parent=sample["Heading3"], fontName="Helvetica-Bold",
            fontSize=11.5, leading=15, textColor=DARK, spaceBefore=3 * mm,
            spaceAfter=1.5 * mm, keepWithNext=True,
        ),
        "body": ParagraphStyle(
            "body", parent=sample["BodyText"], fontName="Helvetica",
            fontSize=9.4, leading=14, textColor=TEXT, alignment=TA_JUSTIFY,
            spaceAfter=2.8 * mm,
        ),
        "body_small": ParagraphStyle(
            "body_small", parent=sample["BodyText"], fontName="Helvetica",
            fontSize=8.2, leading=11.5, textColor=TEXT, alignment=TA_JUSTIFY,
            spaceAfter=1.7 * mm,
        ),
        "muted": ParagraphStyle(
            "muted", parent=sample["BodyText"], fontName="Helvetica",
            fontSize=8, leading=11, textColor=MUTED, alignment=TA_LEFT,
        ),
        "table": ParagraphStyle(
            "table", parent=sample["BodyText"], fontName="Helvetica",
            fontSize=7.6, leading=9.5, textColor=TEXT, alignment=TA_LEFT,
        ),
        "table_head": ParagraphStyle(
            "table_head", parent=sample["BodyText"], fontName="Helvetica-Bold",
            fontSize=7.5, leading=9, textColor=WHITE, alignment=TA_LEFT,
        ),
        "metric": ParagraphStyle(
            "metric", parent=sample["BodyText"], fontName="Helvetica-Bold",
            fontSize=17, leading=20, textColor=NAVY, alignment=TA_CENTER,
        ),
        "metric_label": ParagraphStyle(
            "metric_label", parent=sample["BodyText"], fontName="Helvetica",
            fontSize=7.3, leading=9.3, textColor=MUTED, alignment=TA_CENTER,
        ),
        "callout": ParagraphStyle(
            "callout", parent=sample["BodyText"], fontName="Helvetica",
            fontSize=8.4, leading=12, textColor=TEXT, alignment=TA_JUSTIFY,
        ),
        "footer": ParagraphStyle(
            "footer", parent=sample["BodyText"], fontName="Helvetica",
            fontSize=7, leading=8, textColor=MUTED, alignment=TA_CENTER,
        ),
    }


def p(text: Any, style: ParagraphStyle) -> Paragraph:
    return Paragraph(esc(text), style)


def hp(text: Any, style: ParagraphStyle) -> Paragraph:
    """HTML-t engedő Paragraph, csak saját, kontrollált szöveghez."""
    return Paragraph(str(text), style)


def header_footer(canvas, doc, lang: str, report_date: date) -> None:
    canvas.saveState()
    width, height = A4
    canvas.setStrokeColor(BORDER)
    canvas.setLineWidth(0.5)
    canvas.line(20 * mm, height - 17 * mm, width - 20 * mm, height - 17 * mm)
    canvas.setFillColor(NAVY)
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawString(20 * mm, height - 12.5 * mm, "TÖRÉSVONALAK")
    canvas.setFillColor(BLUE)
    canvas.setFont("Helvetica-Bold", 7)
    canvas.drawString(20 * mm, height - 15.5 * mm, "INTELLIGENCE HUB")
    canvas.setFillColor(MUTED)
    canvas.setFont("Helvetica", 7.5)
    title = "Ukrajnai Fronthelyzet" if lang == "hu" else "Ukraine Frontline"
    canvas.drawRightString(width - 20 * mm, height - 12.5 * mm, title)
    canvas.drawRightString(width - 20 * mm, height - 15.5 * mm, human_date(report_date, lang))

    canvas.line(20 * mm, 15 * mm, width - 20 * mm, 15 * mm)
    canvas.setFont("Helvetica", 7)
    canvas.drawString(20 * mm, 10.5 * mm, f"Törésvonalak Intelligence Hub | v{VERSION}")
    canvas.drawRightString(width - 20 * mm, 10.5 * mm, str(doc.page))
    canvas.restoreState()


def cover_page(canvas, doc, lang: str, report_date: date) -> None:
    canvas.saveState()
    width, height = A4
    canvas.setFillColor(NAVY)
    canvas.rect(0, 0, width, height, stroke=0, fill=1)
    canvas.setFillColor(BLUE)
    canvas.rect(0, 0, width, 42 * mm, stroke=0, fill=1)

    canvas.setFillColor(WHITE)
    canvas.setFont("Helvetica-Bold", 17)
    canvas.drawCentredString(width / 2, height - 45 * mm, "TÖRÉSVONALAK")
    canvas.setFont("Helvetica-Bold", 11)
    canvas.drawCentredString(width / 2, height - 53 * mm, "INTELLIGENCE HUB")
    canvas.setFont("Helvetica", 8.5)
    canvas.setFillColor(colors.HexColor("#D5E5F1"))
    canvas.drawCentredString(
        width / 2,
        height - 60 * mm,
        "Geopolitika • Biztonságpolitika • Ellátásbiztonság • OSINT-elemzés",
    )

    canvas.setStrokeColor(colors.HexColor("#8AB2CF"))
    canvas.setLineWidth(1)
    canvas.line(42 * mm, height - 72 * mm, width - 42 * mm, height - 72 * mm)

    canvas.setFillColor(WHITE)
    canvas.setFont("Helvetica-Bold", 24)
    title_lines = (
        ["UKRAJNAI FRONTHELYZET", "NAPI STRATÉGIAI HÍRSZERZÉSI JELENTÉS"]
        if lang == "hu"
        else ["UKRAINE FRONTLINE", "DAILY STRATEGIC INTELLIGENCE REPORT"]
    )
    y = height - 100 * mm
    for line in title_lines:
        canvas.drawCentredString(width / 2, y, line)
        y -= 10 * mm

    canvas.setFont("Helvetica", 10.5)
    canvas.setFillColor(colors.HexColor("#DCEAF4"))
    subtitle = (
        "Nyílt forrású információk integrált, számszerű és szöveges értékelése"
        if lang == "hu"
        else "Integrated quantitative and narrative assessment of open-source information"
    )
    canvas.drawCentredString(width / 2, y - 4 * mm, subtitle)

    canvas.setFillColor(WHITE)
    canvas.setFont("Helvetica-Bold", 13)
    canvas.drawCentredString(width / 2, 59 * mm, human_date(report_date, lang))
    canvas.setFont("Helvetica", 8.5)
    canvas.setFillColor(colors.HexColor("#DCEAF4"))
    workflow = (
        "Nyílt források  →  szemantikai és térbeli elemzés  →  stratégiai mutatók  →  integrált értékelés"
        if lang == "hu"
        else "Open sources  →  semantic and geospatial analysis  →  strategic indicators  →  integrated assessment"
    )
    canvas.drawCentredString(width / 2, 48 * mm, workflow)
    canvas.restoreState()


def doc_template(output: Path, lang: str, report_date: date) -> BaseDocTemplate:
    output.parent.mkdir(parents=True, exist_ok=True)
    doc = BaseDocTemplate(
        str(output),
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=22 * mm,
        bottomMargin=20 * mm,
        title=("Ukrajnai Fronthelyzet" if lang == "hu" else "Ukraine Frontline"),
        author="Törésvonalak Intelligence Hub",
        subject="Daily strategic intelligence report",
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="main")
    doc.addPageTemplates([
        PageTemplate(id="cover", frames=[frame], onPage=lambda c, d: cover_page(c, d, lang, report_date)),
        PageTemplate(id="body", frames=[frame], onPage=lambda c, d: header_footer(c, d, lang, report_date)),
    ])
    return doc


def freshness_color(status: str) -> colors.Color:
    return {
        "current": GREEN,
        "older": AMBER,
        "stale": RED,
        "missing": MUTED,
    }.get(status, MUTED)


def freshness_label(status: str, lang: str) -> str:
    if lang == "hu":
        return {
            "current": "aktuális",
            "older": "korábbi",
            "stale": "elavult",
            "missing": "ismeretlen",
        }.get(status, status)
    return {
        "current": "current",
        "older": "older",
        "stale": "stale",
        "missing": "unknown",
    }.get(status, status)


def freshness_callout(fresh: Freshness, lang: str, styles: Mapping[str, ParagraphStyle]) -> Table:
    if lang == "hu":
        updated = fresh.updated_at.strftime("%Y-%m-%d") if fresh.updated_at else "-"
        coverage = (
            f"{fresh.period_start.isoformat()} - {fresh.period_end.isoformat()}"
            if fresh.period_start and fresh.period_end
            else "-"
        )
        text = (
            f"<b>ADATGYAKORISÁG:</b> {esc(fresh.frequency.upper())} &nbsp;&nbsp; "
            f"<b>UTOLSÓ FRISSÍTÉS:</b> {updated} &nbsp;&nbsp; "
            f"<b>LEFEDETTSÉG:</b> {coverage}<br/>"
            f"<b>STÁTUSZ:</b> {freshness_label(fresh.status, lang).upper()} - {esc(fresh.note)}"
        )
    else:
        updated = fresh.updated_at.strftime("%Y-%m-%d") if fresh.updated_at else "-"
        coverage = (
            f"{fresh.period_start.isoformat()} - {fresh.period_end.isoformat()}"
            if fresh.period_start and fresh.period_end
            else "-"
        )
        text = (
            f"<b>DATA FREQUENCY:</b> {esc(fresh.frequency.upper())} &nbsp;&nbsp; "
            f"<b>LAST UPDATE:</b> {updated} &nbsp;&nbsp; "
            f"<b>COVERAGE:</b> {coverage}<br/>"
            f"<b>STATUS:</b> {freshness_label(fresh.status, lang).upper()} - {esc(fresh.note)}"
        )
    table = Table([[hp(text, styles["callout"])]], colWidths=[170 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), PALE_BLUE),
        ("BOX", (0, 0), (-1, -1), 0.8, freshness_color(fresh.status)),
        ("LINEBEFORE", (0, 0), (0, -1), 4, freshness_color(fresh.status)),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    return table


def metric_cards(cards: Sequence[tuple[str, str]], styles: Mapping[str, ParagraphStyle]) -> Table:
    rows = []
    for i in range(0, len(cards), 3):
        chunk = list(cards[i:i + 3])
        while len(chunk) < 3:
            chunk.append(("", ""))
        rows.append([
            Table(
                [[p(value, styles["metric"])], [p(label, styles["metric_label"])]],
                colWidths=[52 * mm],
                rowHeights=[12 * mm, 10 * mm],
                style=TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), PALE_BLUE),
                    ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]),
            )
            for value, label in chunk
        ])
    outer = Table(rows, colWidths=[56.5 * mm] * 3, hAlign="LEFT")
    outer.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    return outer


def make_table(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    styles: Mapping[str, ParagraphStyle],
    widths: Sequence[float] | None = None,
    repeat_rows: int = 1,
) -> Table:
    data = [[p(h, styles["table_head"]) for h in headers]]
    for row in rows:
        data.append([p(cell, styles["table"]) for cell in row])
    table = Table(data, colWidths=widths, repeatRows=repeat_rows, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("GRID", (0, 0), (-1, -1), 0.45, BORDER),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, PALE_BLUE]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def report_paths(output_dir: Path, report_date: date, lang: str) -> ReportPaths:
    archive_dir = output_dir / "archive"
    return ReportPaths(
        archive=archive_dir / f"{report_date.isoformat()}-{lang}.pdf",
        latest=output_dir / f"latest-{lang}.pdf",
    )


def executive_text(
    lang: str,
    report_date: date,
    territorial: Mapping[str, Any],
    osint: Mapping[str, Any],
    firms_24: int,
    long_term: Mapping[str, Any],
    deep_fresh: Freshness,
) -> list[str]:
    ru = territorial["ru_gain"]
    ua = territorial["ua_gain"]
    net = territorial["net_ru"]
    active = long_term["active_sector"]
    threat = long_term["threat_sector"]
    threat_level = long_term["threat_level"]

    if lang == "hu":
        if net > 0.05:
            balance = f"Az azonosított napi területi egyenleg Oroszország javára {num(net)} km²."
        elif net < -0.05:
            balance = f"Az azonosított napi területi egyenleg Ukrajna javára {num(abs(net))} km²."
        else:
            balance = "A napi területi egyenleg lényegében változatlan frontvonalat jelez."

        paragraph1 = (
            f"A {human_date(report_date, lang)} dátumú összkép alapján az adatállomány "
            f"{num(ru)} km² orosz területszerzést és {num(ua)} km² ukrán visszafoglalást azonosít. "
            f"{balance} A számok taktikai mozgást mutatnak; önmagukban nem bizonyítanak hadműveleti áttörést."
        )
        paragraph2 = (
            f"A hosszú távú összesítés szerint a legaktívabb térség: <b>{esc(active)}</b>. "
            f"A legmagasabb jelzett fenyegetés a <b>{esc(threat)}</b> szektorhoz kapcsolódik, "
            f"<b>{esc(threat_level)}</b> besorolással. Az értékelés súlyát növeli vagy csökkenti, "
            f"hogy a területi változás, az eseménykoncentráció és a hőpontok ugyanabba az irányba mutatnak-e."
        )
        paragraph3 = (
            f"A napi OSINT-adatfolyam {osint['count']} egyedi elemet, a 24 órás FIRMS-adatállomány "
            f"{firms_24} hőpontot tartalmaz. A FIRMS-adatokat aktivitási jelzésként kezeljük, "
            f"nem automatikus katonai eseményazonosításként."
        )
        if deep_fresh.status == "current":
            paragraph4 = (
                "A mélységi csapások blokkja a legutóbbi teljes heti adatállományt mutatja. "
                "Ez stratégiai háttérként része az integrált értékelésnek, de nem azonosítható "
                "automatikusan a jelentés napjának kizárólagos aktivitásaként."
            )
        elif deep_fresh.status == "older":
            paragraph4 = (
                "A mélységi csapások adatállománya korábbi heti periódust fed le. "
                "A jelentés ezért csak csökkentett súlyú háttérjelzésként használja."
            )
        else:
            paragraph4 = (
                "A mélységi csapások adatállománya elavult vagy dátuma nem értelmezhető. "
                "A napi integrált következtetésbe nem számít bele."
            )
    else:
        if net > 0.05:
            balance = f"The identified daily territorial balance favours Russia by {num(net)} km²."
        elif net < -0.05:
            balance = f"The identified daily territorial balance favours Ukraine by {num(abs(net))} km²."
        else:
            balance = "The daily territorial balance indicates an essentially unchanged frontline."

        paragraph1 = (
            f"The {human_date(report_date, lang)} dataset identifies {num(ru)} km² of Russian gains "
            f"and {num(ua)} km² of Ukrainian recaptures. {balance} These figures indicate tactical "
            f"movement and do not by themselves demonstrate an operational breakthrough."
        )
        paragraph2 = (
            f"The long-term summary identifies <b>{esc(active)}</b> as the most active area. "
            f"The highest indicated threat is associated with <b>{esc(threat)}</b>, rated "
            f"<b>{esc(threat_level)}</b>. Assessment confidence rises when territorial change, "
            f"event concentration and hotspot patterns point in the same direction."
        )
        paragraph3 = (
            f"The daily OSINT feed contains {osint['count']} unique items and the 24-hour FIRMS "
            f"dataset contains {firms_24} hotspots. FIRMS observations are treated as activity "
            f"indicators rather than automatic confirmation of military events."
        )
        if deep_fresh.status == "current":
            paragraph4 = (
                "The deep-strike section presents the latest complete weekly dataset. It informs "
                "the integrated assessment as strategic context but is not presented as activity "
                "occurring exclusively on the report date."
            )
        elif deep_fresh.status == "older":
            paragraph4 = (
                "The deep-strike dataset covers an older weekly period and is therefore used only "
                "as reduced-weight contextual evidence."
            )
        else:
            paragraph4 = (
                "The deep-strike dataset is stale or lacks an interpretable date and is excluded "
                "from the integrated daily conclusion."
            )
    return [paragraph1, paragraph2, paragraph3, paragraph4]


def integrated_assessment(
    lang: str,
    territorial: Mapping[str, Any],
    osint: Mapping[str, Any],
    firms_24: int,
    deep_count: int,
    deep_fresh: Freshness,
    units: Mapping[str, Any],
) -> tuple[str, list[str]]:
    score = 0.0
    reasons: list[str] = []

    net = territorial["net_ru"]
    if net > 0.25:
        score += 2
        reasons.append("orosz területi előny" if lang == "hu" else "Russian territorial advantage")
    elif net > 0.05:
        score += 1
        reasons.append("korlátozott orosz területi előny" if lang == "hu" else "limited Russian territorial advantage")
    elif net < -0.25:
        score -= 2
        reasons.append("ukrán területi előny" if lang == "hu" else "Ukrainian territorial advantage")
    elif net < -0.05:
        score -= 1
        reasons.append("korlátozott ukrán területi előny" if lang == "hu" else "limited Ukrainian territorial advantage")

    if osint["count"] >= 40:
        score += 0.5
        reasons.append("magas eseménysűrűség" if lang == "hu" else "high event density")
    if firms_24 >= 100:
        score += 0.5
        reasons.append("magas 24 órás hőpontszám" if lang == "hu" else "high 24-hour hotspot count")
    if units["count"] == 0:
        reasons.append("korlátozott egységadat-lefedettség" if lang == "hu" else "limited unit-data coverage")
    if deep_fresh.include_in_daily and deep_count:
        score += min(0.5, deep_count / 100)
        reasons.append("aktuális heti mélységi aktivitás" if lang == "hu" else "current weekly deep-strike activity")

    if score >= 1.5:
        label = "Orosz taktikai momentum" if lang == "hu" else "Russian tactical momentum"
    elif score <= -1.5:
        label = "Ukrán taktikai momentum" if lang == "hu" else "Ukrainian tactical momentum"
    else:
        label = "Kiegyensúlyozott vagy lokálisan megosztott helyzet" if lang == "hu" else "Balanced or locally fragmented situation"
    return label, reasons


def build_pdf(datasets: Mapping[str, Any], lang: str, output: Path, report_date: date) -> None:
    tr = i18n(lang)
    styles = make_styles()

    territorial = territorial_metrics(datasets["territorial_daily"])
    territorial_30 = territorial_metrics(datasets["territorial_30"])
    osint = osint_metrics(datasets["osint"], datasets["telegram"])
    units = unit_metrics(datasets["units"])
    long_term = long_term_metrics(datasets["long_term"])
    firms = {days: firms_points(datasets[f"firms_{days}"]) for days in (1, 3, 10, 30)}
    deep_items = extract_items(datasets["deep_strikes"])

    freshness = {
        "territorial_daily": freshness_for(
            datasets["territorial_daily"], INPUTS["territorial_daily"], report_date,
            tr["daily"], 2, 4, lang,
        ),
        "long_term": freshness_for(
            datasets["long_term"], INPUTS["long_term"], report_date,
            tr["daily"], 2, 5, lang,
        ),
        "osint": freshness_for(
            datasets["osint"], INPUTS["osint"], report_date,
            tr["daily"], 2, 5, lang,
        ),
        "firms_1": freshness_for(
            datasets["firms_1"], INPUTS["firms_1"], report_date,
            tr["daily"], 2, 4, lang,
        ),
        "firms_30": freshness_for(
            datasets["firms_30"], INPUTS["firms_30"], report_date,
            tr["rolling"], 3, 7, lang,
        ),
        "units": freshness_for(
            datasets["units"], INPUTS["units"], report_date,
            tr["periodic"], 7, 21, lang,
        ),
        "deep": freshness_for(
            datasets["deep_strikes"], INPUTS["deep_strikes"], report_date,
            tr["weekly"], 7, 14, lang,
        ),
    }

    doc = doc_template(output, lang, report_date)
    story: list[Any] = [PageBreak()]
    doc.handle_nextPageTemplate("body")

    story += [
        p(tr["about"], styles["h1"]),
        p(tr["about_text"], styles["body"]),
        Spacer(1, 2 * mm),
        p(tr["input_models"], styles["h1"]),
    ]

    source_rows = []
    source_specs = [
        ("Területi változás" if lang == "hu" else "Territorial change", freshness["territorial_daily"], INPUTS["territorial_daily"]),
        ("Hosszú távú összesítés" if lang == "hu" else "Long-term summary", freshness["long_term"], INPUTS["long_term"]),
        ("OSINT-hírek" if lang == "hu" else "OSINT reporting", freshness["osint"], INPUTS["osint"]),
        ("FIRMS 24 óra" if lang == "hu" else "FIRMS 24 hours", freshness["firms_1"], INPUTS["firms_1"]),
        ("FIRMS 30 nap" if lang == "hu" else "FIRMS 30 days", freshness["firms_30"], INPUTS["firms_30"]),
        ("Egységadatok" if lang == "hu" else "Unit observations", freshness["units"], INPUTS["units"]),
        ("Mélységi csapások" if lang == "hu" else "Deep strikes", freshness["deep"], INPUTS["deep_strikes"]),
    ]
    for name, fresh, path in source_specs:
        updated = fresh.updated_at.strftime("%Y-%m-%d") if fresh.updated_at else "-"
        source_rows.append([
            name,
            fresh.frequency,
            updated,
            freshness_label(fresh.status, lang),
            str(path.relative_to(ROOT)),
        ])
    story += [
        make_table(
            ["Adatblokk", "Gyakoriság", "Frissítve", "Státusz", "Fájl"]
            if lang == "hu"
            else ["Data block", "Frequency", "Updated", "Status", "File"],
            source_rows,
            styles,
            widths=[35 * mm, 22 * mm, 22 * mm, 21 * mm, 70 * mm],
        ),
        Spacer(1, 4 * mm),
        p(tr["glance"], styles["h1"]),
    ]

    integrated_label, integrated_reasons = integrated_assessment(
        lang, territorial, osint, len(firms[1]), len(deep_items), freshness["deep"], units
    )
    cards = [
        (f"{num(territorial['ru_gain'])} km²", "Orosz napi területszerzés" if lang == "hu" else "Russian daily gains"),
        (f"{num(territorial['ua_gain'])} km²", "Ukrán napi visszafoglalás" if lang == "hu" else "Ukrainian daily recaptures"),
        (f"{num(territorial['net_ru'], signed=True)} km²", "Nettó változás Oroszország javára" if lang == "hu" else "Net change in Russia's favour"),
        (str(osint["count"]), "OSINT-elemek" if lang == "hu" else "OSINT items"),
        (str(len(firms[1])), "FIRMS-hőpont, 24 óra" if lang == "hu" else "FIRMS hotspots, 24h"),
        (str(len(deep_items)), "Heti mélységi rekord" if lang == "hu" else "Weekly deep-strike records"),
        (str(long_term["active_sector"]), "Legaktívabb szektor" if lang == "hu" else "Most active sector"),
        (str(long_term["threat_level"]), "Legmagasabb fenyegetési szint" if lang == "hu" else "Highest threat level"),
        (integrated_label, "Integrált napi értékelés" if lang == "hu" else "Integrated daily assessment"),
    ]
    story += [metric_cards(cards, styles), Spacer(1, 4 * mm)]

    story += [p(tr["executive"], styles["h1"])]
    for paragraph in executive_text(
        lang, report_date, territorial, osint, len(firms[1]), long_term, freshness["deep"]
    ):
        story.append(hp(paragraph, styles["body"]))

    story += [
        p(tr["territory"], styles["h1"]),
        freshness_callout(freshness["territorial_daily"], lang, styles),
        Spacer(1, 3 * mm),
    ]
    territory_rows = [
        ["1 nap" if lang == "hu" else "1 day", num(territorial["ru_gain"]), num(territorial["ua_gain"]), num(territorial["net_ru"], signed=True)],
        ["30 nap" if lang == "hu" else "30 days", num(territorial_30["ru_gain"]), num(territorial_30["ua_gain"]), num(territorial_30["net_ru"], signed=True)],
    ]
    story += [
        make_table(
            ["Időszak", "Orosz nyereség (km²)", "Ukrán visszafoglalás (km²)", "Nettó RU (km²)"]
            if lang == "hu"
            else ["Period", "Russian gains (km²)", "Ukrainian recaptures (km²)", "Net RU (km²)"],
            territory_rows,
            styles,
            widths=[35 * mm, 43 * mm, 48 * mm, 44 * mm],
        ),
        Spacer(1, 3 * mm),
    ]

    top_changes = territorial["rows"][:10]
    if top_changes:
        rows = [
            [
                item["actor"],
                num(item["area"]),
                str(item["sector"]),
                str(item["place"]),
            ]
            for item in top_changes
        ]
        story += [
            p("Legfontosabb azonosított változások" if lang == "hu" else "Largest identified changes", styles["h2"]),
            make_table(
                ["Fél", "Terület (km²)", "Szektor", "Helyszín"]
                if lang == "hu"
                else ["Side", "Area (km²)", "Sector", "Location"],
                rows,
                styles,
                widths=[16 * mm, 28 * mm, 48 * mm, 78 * mm],
            ),
        ]
    else:
        story.append(p(
            "A napi területi GeoJSON nem tartalmazott külön értelmezhető változási poligonokat."
            if lang == "hu"
            else "The daily territorial GeoJSON contained no individually interpretable change polygons.",
            styles["body"],
        ))

    if lang == "hu":
        territory_analysis = (
            f"A napi területi változás {num(territorial['feature_count'], 0)} elkülönített elemre épül. "
            f"A legnagyobb változások helyszíneit önmagukban nem tekintjük áttörésnek. Hadműveleti jelentőség "
            f"akkor valószínűsíthető, ha több egymást követő napon azonos irányú területi mozgás, magas "
            f"frontaktivitás és logisztikai vagy mélységi nyomás is megjelenik."
        )
    else:
        territory_analysis = (
            f"The daily territorial assessment is based on {territorial['feature_count']} separate features. "
            f"The largest changes are not automatically classified as breakthroughs. Operational significance "
            f"requires persistent movement in the same direction, elevated frontline activity and supporting "
            f"logistical or deep-strike pressure."
        )
    story += [p(territory_analysis, styles["body"])]

    story += [
        p(tr["activity"], styles["h1"]),
        freshness_callout(freshness["long_term"], lang, styles),
        Spacer(1, 3 * mm),
    ]
    activity_rows = [
        ["Legaktívabb szektor" if lang == "hu" else "Most active sector", long_term["active_sector"]],
        ["Legmagasabb fenyegetés szektora" if lang == "hu" else "Highest-threat sector", long_term["threat_sector"]],
        ["Fenyegetési szint" if lang == "hu" else "Threat level", long_term["threat_level"]],
        ["Konfliktusindex" if lang == "hu" else "Conflict index", num(long_term["conflict_index"])],
        ["Napi események" if lang == "hu" else "Daily events", str(long_term["daily_events"] or osint["count"])],
    ]
    story += [
        make_table(
            ["Mutató", "Aktuális érték"] if lang == "hu" else ["Indicator", "Current value"],
            activity_rows,
            styles,
            widths=[78 * mm, 92 * mm],
        ),
        Spacer(1, 3 * mm),
    ]

    top_sectors = osint["sector_counts"].most_common(8)
    if top_sectors:
        rows = [[sector, count, f"{count / max(1, osint['count']) * 100:.1f}%"] for sector, count in top_sectors]
        story += [
            p("Eseménykoncentráció szektoronként" if lang == "hu" else "Event concentration by sector", styles["h2"]),
            make_table(
                ["Szektor", "Esemény", "Arány"] if lang == "hu" else ["Sector", "Events", "Share"],
                rows,
                styles,
                widths=[95 * mm, 35 * mm, 40 * mm],
            ),
        ]

    story += [
        p(tr["firms"], styles["h1"]),
        freshness_callout(freshness["firms_1"], lang, styles),
        Spacer(1, 3 * mm),
    ]
    firms_rows = [
        ["24 óra" if lang == "hu" else "24 hours", len(firms[1])],
        ["3 nap" if lang == "hu" else "3 days", len(firms[3])],
        ["10 nap" if lang == "hu" else "10 days", len(firms[10])],
        ["30 nap" if lang == "hu" else "30 days", len(firms[30])],
    ]
    story += [
        make_table(
            ["Időablak", "Azonosított hőpontok"] if lang == "hu" else ["Time window", "Identified hotspots"],
            firms_rows,
            styles,
            widths=[80 * mm, 90 * mm],
        ),
        Spacer(1, 3 * mm),
    ]
    if lang == "hu":
        firms_text = (
            "A FIRMS-adatok termikus anomáliákat jeleznek. A hőpontok származhatnak harci eseményből, "
            "ipari tevékenységből, mezőgazdasági égetésből vagy más hőforrásból. A jelentés ezért csak "
            "akkor emeli a hőpontok elemzési súlyát, ha térben és időben egybeesnek megerősített "
            "fronteseményekkel, területi változással vagy hiteles OSINT-jelentésekkel."
        )
    else:
        firms_text = (
            "FIRMS data indicate thermal anomalies. Hotspots may originate from combat, industrial "
            "activity, agricultural burning or other heat sources. Their analytical weight rises only "
            "when they coincide spatially and temporally with confirmed frontline events, territorial "
            "change or credible OSINT reporting."
        )
    story.append(p(firms_text, styles["body"]))

    story += [
        p(tr["deep"], styles["h1"]),
        freshness_callout(freshness["deep"], lang, styles),
        Spacer(1, 3 * mm),
    ]
    if deep_items:
        deep_rows = []
        for row in deep_items[-15:]:
            keys = {str(k).lower(): v for k, v in row.items()}
            date_value = next((v for k, v in keys.items() if "date" in k or "dátum" in k or "datum" in k), "-")
            place = next((v for k, v in keys.items() if any(t in k for t in ("hely", "location", "települ", "celpont", "célpont"))), "-")
            attack = next((v for k, v in keys.items() if any(t in k for t in ("type", "típus", "tipus", "jelleg"))), "-")
            desc = next((v for k, v in keys.items() if any(t in k for t in ("leírás", "leiras", "description", "summary", "esemény"))), "-")
            deep_rows.append([str(date_value)[:10], place, attack, str(desc)[:250]])
        story += [
            make_table(
                ["Dátum", "Helyszín/célpont", "Típus", "Rövid leírás"]
                if lang == "hu"
                else ["Date", "Location/target", "Type", "Brief description"],
                deep_rows,
                styles,
                widths=[22 * mm, 43 * mm, 32 * mm, 73 * mm],
            ),
            Spacer(1, 3 * mm),
        ]
        if lang == "hu":
            deep_text = (
                f"A blokk {len(deep_items)} rekordot tartalmaz a legutóbbi elérhető heti állományból. "
                f"Az adatok nem kizárólag a jelentés napját írják le. Az integrált napi értékelésben "
                f"{'teljes háttérsúllyal' if freshness['deep'].status == 'current' else 'csökkentett súllyal' if freshness['deep'].status == 'older' else 'nem'} szerepelnek."
            )
        else:
            deep_text = (
                f"The section contains {len(deep_items)} records from the latest available weekly dataset. "
                f"They do not describe the report date exclusively and are "
                f"{'used as current contextual evidence' if freshness['deep'].status == 'current' else 'used with reduced weight' if freshness['deep'].status == 'older' else 'excluded from the daily integrated assessment'}."
            )
        story.append(p(deep_text, styles["body"]))
    else:
        story.append(p(
            "A heti mélységi csapásokat tartalmazó Excel nem elérhető vagy nem olvasható."
            if lang == "hu"
            else "The weekly deep-strike Excel file is unavailable or unreadable.",
            styles["body"],
        ))

    story += [
        p(tr["units"], styles["h1"]),
        freshness_callout(freshness["units"], lang, styles),
        Spacer(1, 3 * mm),
    ]
    if units["count"]:
        side_rows = [[side, count] for side, count in units["sides"].most_common()]
        story += [
            make_table(
                ["Fél / besorolás", "Azonosított rekordok"] if lang == "hu" else ["Side / classification", "Identified records"],
                side_rows,
                styles,
                widths=[105 * mm, 65 * mm],
            ),
            Spacer(1, 3 * mm),
        ]
        top_unit_sectors = units["sectors"].most_common(10)
        story += [
            make_table(
                ["Szektor", "Egységrekordok"] if lang == "hu" else ["Sector", "Unit records"],
                [[name, count] for name, count in top_unit_sectors],
                styles,
                widths=[105 * mm, 65 * mm],
            ),
        ]
    else:
        story.append(p(
            "Az egységadat-fájl nem tartalmazott értelmezhető rekordokat. Ez nem jelenti azt, hogy az adott napon nem történt átcsoportosítás."
            if lang == "hu"
            else "The unit-data file contained no interpretable records. This does not mean that no force movements occurred.",
            styles["body"],
        ))

    story += [
        p(tr["osint"], styles["h1"]),
        freshness_callout(freshness["osint"], lang, styles),
        Spacer(1, 3 * mm),
    ]
    top_events = osint["items"][:12]
    if top_events:
        event_rows = []
        for item in top_events:
            event_rows.append([
                str(item.get("date") or item.get("published_at") or "-")[:10],
                first_value(item, (("sectorShortName",), ("sectorName",), ("sector",)), default="-"),
                first_value(item, (("category",), ("type",), ("event_type",)), default="-"),
                first_value(item, (("title",), ("summary",), ("description",)), default="-"),
                first_value(item, (("sourceName",), ("source",), ("sourceType",)), default="-"),
            ])
        story += [
            make_table(
                ["Dátum", "Szektor", "Típus", "Esemény", "Forrás"]
                if lang == "hu"
                else ["Date", "Sector", "Type", "Event", "Source"],
                event_rows,
                styles,
                widths=[19 * mm, 34 * mm, 25 * mm, 67 * mm, 25 * mm],
            )
        ]
    else:
        story.append(p(
            "A napi OSINT-adatfolyam nem tartalmazott értelmezhető eseményeket."
            if lang == "hu"
            else "The daily OSINT feed contained no interpretable events.",
            styles["body"],
        ))

    story += [p(tr["integrated"], styles["h1"])]
    if lang == "hu":
        integrated_text = (
            f"Az integrált minősítés: <b>{esc(integrated_label)}</b>. A besorolást elsősorban a napi "
            f"területi egyenleg, az eseménysűrűség, a 24 órás FIRMS-aktivitás, az egységadatok "
            f"lefedettsége és - ha megfelelően friss - a legutóbbi heti mélységi aktivitás alakítja. "
            f"A fő alátámasztó tényezők: {esc(', '.join(integrated_reasons) if integrated_reasons else 'nincs egyértelmű domináns tényező')}."
        )
    else:
        integrated_text = (
            f"The integrated classification is <b>{esc(integrated_label)}</b>. It is driven primarily by "
            f"the daily territorial balance, event density, 24-hour FIRMS activity, unit-data coverage "
            f"and - when sufficiently current - the latest weekly deep-strike activity. Main supporting "
            f"factors: {esc(', '.join(integrated_reasons) if integrated_reasons else 'no single dominant factor')}."
        )
    story.append(hp(integrated_text, styles["body"]))

    if lang == "hu":
        outlook_points = [
            "A lokális területi mozgást csak akkor érdemes hadműveleti áttörésként kezelni, ha több egymást követő frissítésben fennmarad.",
            "A legaktívabb és a legmagasabb fenyegetésű szektorban figyelni kell az események térbeli sűrűsödését és a szomszédos frontszakaszokra gyakorolt nyomást.",
            "A FIRMS-hőpontok hirtelen emelkedése önmagában nem döntő; a megerősített csapás- és frontadatokkal való egybeesés a lényeges.",
            "A heti mélységi csapások következő frissítése módosíthatja a logisztikai és hátországi nyomásról alkotott képet.",
        ]
    else:
        outlook_points = [
            "Local territorial movement should be treated as an operational breakthrough only if it persists across successive updates.",
            "In the most active and highest-threat sectors, watch for spatial concentration of events and pressure spreading to adjacent sectors.",
            "A sudden rise in FIRMS hotspots is not decisive by itself; coincidence with confirmed strike and frontline reporting is what matters.",
            "The next weekly deep-strike update may alter the assessment of logistical and rear-area pressure.",
        ]
    story += [
        p(tr["outlook"], styles["h2"]),
        *[hp(f"• {esc(point)}", styles["body"]) for point in outlook_points],
    ]

    story += [
        p(tr["method"], styles["h1"]),
    ]
    if lang == "hu":
        method_sections = [
            ("1. Adatgyűjtés", "A jelentés kizárólag a repóban elérhető, nyílt forrásból származó adatállományokat használja. A hiányzó fájl nem kerül automatikusan nullaként értelmezésre."),
            ("2. Adatfrissesség", "Minden blokk saját frissítési gyakorisággal rendelkezik. A napi adatok rövid toleranciát kapnak, a heti mélységi adatok 0-7 nap között aktuálisak, 8-14 nap között csökkentett súlyúak, 14 nap felett elavultak."),
            ("3. Területi változás", "A GeoJSON-poligonokból összesített km²-adatok taktikai területi mozgást jeleznek. A geometriai pontosság és a forrásfrissítés időpontja korlátozhatja az eredményt."),
            ("4. OSINT és eseménysűrűség", "Az események deduplikálása URL vagy cím alapján történik. A mennyiségi sűrűség nem egyenlő a katonai jelentőséggel, ezért a jelentés kiemeli a magasabb fontosságú és térben koncentrált elemeket."),
            ("5. FIRMS", "A termikus anomáliák nem automatikusan katonai események. Csak más adatforrásokkal együtt növelik az értékelés bizonyosságát."),
            ("6. Egységadatok", "Az egységadatok az utolsó ismert megfigyeléseket mutatják. Az adathiány nem bizonyítja az egység hiányát vagy mozdulatlanságát."),
            ("7. Mélységi csapások", "A data/manual/ukrajna_oroszorszag_melysegi_csapasok.xlsx heti frissítésű. A PDF mindig feltünteti a lefedett időszakot, és nem nevezi ezeket automatikusan napi eseményeknek."),
            ("8. Integrált minősítés", "Az összegző besorolás jelzőrendszer, nem valószínűségi modell. A napi területi egyenleget, eseménysűrűséget, hőpontokat, egységadatokat és friss heti háttéradatokat kombinálja."),
            ("9. Bizonytalanság", "A fronthelyzet gyorsan változhat. A források késhetnek, egymásnak ellentmondhatnak vagy részleges képet adhatnak. A jelentés ezért nem helyettesíti az elsődleges katonai hírszerzést."),
        ]
    else:
        method_sections = [
            ("1. Data collection", "The report uses only open-source datasets available in the repository. A missing file is not automatically interpreted as a zero value."),
            ("2. Data freshness", "Each block has its own update frequency. Daily data receive a short tolerance; weekly deep-strike data are current at 0-7 days, reduced-weight at 8-14 days and stale beyond 14 days."),
            ("3. Territorial change", "Aggregated km² values derived from GeoJSON polygons indicate tactical territorial movement. Geometric accuracy and source timing constrain the result."),
            ("4. OSINT and event density", "Events are deduplicated by URL or title. Quantity does not equal military significance, so the report prioritises higher-importance and spatially concentrated observations."),
            ("5. FIRMS", "Thermal anomalies are not automatically military events. They increase confidence only when supported by other data sources."),
            ("6. Unit observations", "Unit data represent the last known observations. Missing data do not prove that a unit is absent or static."),
            ("7. Deep strikes", "data/manual/ukrajna_oroszorszag_melysegi_csapasok.xlsx is updated weekly. The PDF displays its coverage period and never automatically labels all records as same-day events."),
            ("8. Integrated classification", "The summary classification is an indicator framework rather than a probability model. It combines daily territorial balance, event density, hotspots, unit observations and sufficiently current weekly context."),
            ("9. Uncertainty", "The frontline can change rapidly. Sources may be delayed, contradictory or incomplete. This report does not replace primary military intelligence."),
        ]
    for heading, body in method_sections:
        story += [p(heading, styles["h3"]), p(body, styles["body"])]

    story += [
        p(tr["disclaimer"], styles["h1"]),
        p(
            (
                "A jelentés nyílt forrású elemzési termék. Nem minősül katonai, befektetési, jogi vagy hivatalos kormányzati tanácsnak. "
                "A szerző és a Törésvonalak nem garantálja az egyes források teljességét vagy hibamentességét."
                if lang == "hu"
                else
                "This report is an open-source analytical product. It is not military, investment, legal or official government advice. "
                "The author and Törésvonalak do not guarantee that individual sources are complete or error-free."
            ),
            styles["body"],
        ),
    ]

    doc.build(story)


def update_index(output_dir: Path) -> None:
    archive_dir = output_dir / "archive"
    archive_dir.mkdir(parents=True, exist_ok=True)
    entries = []
    pattern = re.compile(r"^(\d{4}-\d{2}-\d{2})-(hu|en)\.pdf$")
    for path in sorted(archive_dir.glob("*.pdf"), reverse=True):
        match = pattern.match(path.name)
        if not match:
            continue
        entries.append({
            "date": match.group(1),
            "language": match.group(2),
            "file": f"archive/{path.name}",
            "size_bytes": path.stat().st_size,
            "updated_at": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat(),
        })

    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "report_type": "Ukraine Frontline Daily Strategic Intelligence Report",
        "version": VERSION,
        "reports": entries,
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    with (output_dir / "reports_index.json").open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def main() -> int:
    args = parse_args()
    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = ROOT / output_dir

    datasets = load_datasets()
    report_date = infer_report_date(datasets, args.report_date)
    languages = ("hu", "en") if args.lang == "all" else (args.lang,)

    print(f"Report date: {report_date.isoformat()}")
    print(f"Selected language: {args.lang}")
    print("Archive policy: always rebuild and replace the report for the selected date.")

    for lang in languages:
        paths = report_paths(output_dir, report_date, lang)
        paths.archive.parent.mkdir(parents=True, exist_ok=True)

        build_pdf(datasets, lang, paths.archive, report_date)
        print(f"Created or refreshed archive: {paths.archive.relative_to(ROOT)}")

        shutil.copy2(paths.archive, paths.latest)
        print(f"Updated latest: {paths.latest.relative_to(ROOT)}")

    update_index(output_dir)
    print(f"Updated archive index: {(output_dir / 'reports_index.json').relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
